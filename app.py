import random
import getopt, sys
import datetime
import os.path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill, Border, Side
def leer(nombre_archivo):
    # Nombre del archivo de texto
    #nombre_archivo = "agosto_sin_repetidos.txt"

    # Crear una lista para almacenar los números
    numeros = []
    repetidos = []

    # Leer los números desde el archivo de texto y eliminar duplicados
    with open(nombre_archivo, 'r') as archivo:
        for linea in archivo:
            if len(linea.strip()) == 0 or linea.strip() == "\n":
                continue
            if not linea.strip().isnumeric():
                continue
            numero = int(linea.strip())
            if numero not in numeros:
                numeros.append(numero)
            else:
                repetidos.append(numero)
    if len(repetidos) > 0:
        print("Existen ID set repetidos", repetidos)
        archivo_salida = f"repetidos_({nombre_archivo})"
        escribir(repetidos, archivo_salida)
        #for numero in repetidos:
            # print(numero)
    return numeros


def aleatorio(numeros, archivo_salida, periodo):
    inicio = 0
    largo = len(numeros)
    if largo == 0:
        print("No se encontraron datos para procesar")
        sys.exit(2)
    fin = largo - 1
    indices = list(range(inicio, fin))
    muestras = 0
    letra_cod = ""
    # determinar cantidad de muestras y letra cod
    if (largo >= 2 and largo <=8):
        letra_cod = "A"
        muestras = 2
    elif (largo >= 9 and largo <= 15):
        letra_cod = "B"
        muestras = 3
    elif (largo >= 16 and largo <= 25):
        letra_cod = "C"
        muestras = 5
    elif (largo >= 26 and largo <= 50):
        letra_cod = "D"
        muestras = 8
    elif (largo >= 51 and largo <= 90):
        letra_cod = "E"
        muestras = 13
    elif (largo >= 91 and largo <= 150):
        letra_cod = "F"
        muestras = 20
    elif (largo >= 151 and largo <= 280):
        letra_cod = "G"
        muestras = 32
    elif (largo >= 281 and largo <= 500):
        letra_cod = "H"
        muestras = 50
    else:
        letra_cod = "I"
        muestras = 60

    random.shuffle(indices)
    # Generar el nombre del archivo con la fecha y hora
    nombre_archivo = f"aleatorios_({archivo_salida})"

    indices_imprimir = indices[:muestras]
    new_numeros = []
    for indice in indices_imprimir:
        new_numeros.append(numeros[indice])

    escribir(new_numeros, nombre_archivo)
    genera_excel(new_numeros, nombre_archivo, largo, letra_cod, periodo)

    return random.randint(inicio, fin)

def escribir(numeros, archivo_salida):
    fecha_hora_actual = datetime.datetime.now()
    formato_fecha_hora = fecha_hora_actual.strftime("%Y-%m-%d_%H-%M-%S")  # Formato: AAAA-MM-DD_HH-MM-SS
    nombre_archivo = f"{archivo_salida}_{formato_fecha_hora}.txt"
    with open(nombre_archivo, 'w') as archivo:
        for numero in numeros:
            archivo.write(str(numero) + "\n")
    print("Se ha escrito la información en el archivo:", nombre_archivo)


def genera_excel(new_numeros, nombre_archivo, lote, letra_cod, periodo):
    # abrir libro excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos aleatorios"
    ws['A1'] = "Muestra"
    for i in range(len(new_numeros)):
        ws.cell(row=i+2, column=1).value = i+1
        ws.cell(row=i+2, column=2).value = new_numeros[i]
    # Guardar el libro
    wb.save(nombre_archivo)
    """
    muestras = len(new_numeros)
    wb = openpyxl.load_workbook("formato.xl")
    ws = wb["resultado"]
    ws.cell(row=11, column=12).value = muestras
    ws.cell(row=11, column=11).value = lote
    ws.cell(row=14, column=11).value = letra_cod
    ws.cell(row=17, column=11).value = periodo
    columna = 1
    for i in range(len(new_numeros)):
        if i >= 15 and i <= 29:
            columna = 2
        elif i >= 30 and i <= 44:
            columna = 3
        elif i >= 45 and i <= 59:
            columna = 4
        fila = i - (columna - 1) * 15
        ws.cell(row=fila+9, column=columna*2 + 1).value = i+1
        ws.cell(row=fila+9, column=columna*2 + 1).value = new_numeros[i]
    wb.save(nombre_archivo + ".xlsx")


def lee_excel(nombre_archivo, mes_filtro, year_filtro, hoja_entrada):
    wb = openpyxl.load_workbook(nombre_archivo)
    if hoja_entrada == None or hoja_entrada == "":
        hoja = "Hoja1"
    else:
        hoja = hoja_entrada
    try:
        ws = wb[hoja]
    except:
        print("No existe la hoja:", hoja)
        sys.exit(2)
    ## En caso de no especificar hoja, se puede usar active => ws = wb.active
    salir = False
    cont = 2
    numeros = []
    repetidos = []
    print("Leyendo el archivo", nombre_archivo, mes_filtro, year_filtro)
    while not salir:
        if ws.cell(row=cont, column=1).value == None and ws.cell(row=cont, column=2).value == None:
            print("Fin del archivo")
            salir = True
        else:
            if ws.cell(row=cont, column=2).value != None:
                ## validar fecha
                if validar_fecha(str(ws.cell(row=cont, column=2).value)):
                    fecha_valida = ws.cell(row=cont, column=2).value
                    ## print("validada", fecha_valida, fecha_valida.month, fecha_valida.year)
                    if fecha_valida.month == int(mes_filtro) and fecha_valida.year == int(year_filtro):
                        ## print("Fecha valida pasa filtro")
                        if str(ws.cell(row=cont, column=1).value).isnumeric():
                            numero = int(str(ws.cell(row=cont, column=1).value))
                            if numero not in numeros:
                                numeros.append(numero)
                            else:
                                repetidos.append(numero)
            cont += 1
    if len(repetidos) > 0:
        print("Existen ID set repetidos", repetidos)
        archivo_salida = f"repetidos_({nombre_archivo})"
        escribir(repetidos, archivo_salida)
    return numeros


def validar_fecha(fecha):
    fecha_validar = fecha[0:10].strip()
    date_format = '%Y-%m-%d'
    try:
        # formatting the date using strptime() function
        date_object = datetime.datetime.strptime(fecha_validar, date_format)
        return True
    # If the date validation goes wrong
    except ValueError:
        # printing the appropriate text if ValueError occurs
        return False


def main(argv):
    # Definir los valores iniciales de los argumentos
    archivo_entrada = ""
    excel_entrada = ""
    hoja_entrada = ""
    cantidad_numerica = 0
    mes = 0
    year = 0

    try:
        # Obtener los argumentos de la línea de comandos
        opts, args = getopt.getopt(argv, "hi:n:m:y:x:s:", ["input=", "numero=", "mes=", "year=", "excel=", "hoja="])
    except getopt.GetoptError:
        print("Hay 2 modos de uso correcto. \n \
            Modo de uso 1: genera_aleatorios -x <archivo_excel> -m <mes a procesar> -y <año a procesar> -s <OPCIONAL hoja excel> \n \
            Modo de uso 2: genera_aleatorios -i <archivo_texto>")
        sys.exit(2)

    # Procesar los argumentos obtenidos
    for opt, arg in opts:
        if opt == '-h':
            print("Hay 2 modos de uso correcto. \n \
            Modo de uso 1: genera_aleatorios -x <archivo_excel> -m <mes a procesar> -y <año a procesar> -s <OPCIONAL hoja excel> \n \
            Modo de uso 2: genera_aleatorios -i <archivo_texto>")
            sys.exit()
        elif opt in ("-i", "--input"):
            archivo_entrada = arg
        elif opt in ("-m", "--mes"):
            mes = arg
        elif opt in ("-y", "--year"):
            year = arg
        elif opt in ("-x", "--excel"):
            excel_entrada = arg
        elif opt in ("-s", "--hoja"):
            hoja_entrada = arg
        elif opt in ("-n", "--numero"):
            try:
                cantidad_numerica = int(arg)
                if cantidad_numerica <= 0:
                    raise ValueError
            except ValueError:
                print("El segundo argumento debe ser un número entero mayor a cero.")
                sys.exit(2)

    if excel_entrada:
        ## verificar que se hayan proporcionado todos los argumentos
        if not year or not mes:
            print("Error en argumentos. Debe proporcionar el mes y año. \n \
            Uso: genera_aleatorios -x <archivo_excel> -m <mes a procesar> -y <año a procesar> -s <OPCIONAL hoja excel>")
            sys.exit(2)
        if not os.path.isfile(excel_entrada):
            print("El archivo de entrada no existe.")
            sys.exit(2)
        if mes not in ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]:
            print("El mes debe ser un número entre 1 y 12.")
            sys.exit(2)
        if year not in ["2022", "2023", "2024", "2025", "2026", "2027", "2028", "2029", "2030"]:
            print("El año debe ser un número entre 2022 y 2030.")
            sys.exit(2)
        print("Archivo de entrada:", excel_entrada, mes, year)
        numeros = lee_excel(excel_entrada, mes, year, hoja_entrada)
        aleatorio(numeros, excel_entrada, "01/"+mes+"/"+year)
    else:
        ## Si no proporciona el excel verificar que se haya proporcionado algún archiuvo de texto
        # Verificar que se hayan proporcionado todos los argumentos
        if not archivo_entrada:
            print("Debe especificar el archivo a procesar. \n \
            Modo de uso 1: genera_aleatorios -x <archivo_excel> -m <mes a procesar> -y <año a procesar> -s <OPCIONAL hoja excel> \n \
            Modo de uso 2: genera_aleatorios -i <archivo_texto>")
            sys.exit(2)
        if not os.path.isfile(archivo_entrada):
            print("El archivo de entrada no existe.")
            sys.exit(2)
        print("Archivo de entrada:", archivo_entrada)
        numeros = leer(archivo_entrada)
        aleatorio(numeros, archivo_entrada, cantidad_numerica)


if __name__ == "__main__":
    main(sys.argv[1:])